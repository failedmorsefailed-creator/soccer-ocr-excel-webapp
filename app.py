import os
import re
import io
from pathlib import Path
from datetime import datetime
from flask import Flask, request, render_template, redirect, url_for, flash, send_file
from werkzeug.utils import secure_filename
from PIL import Image, ImageOps, ImageFilter
import pytesseract
import pandas as pd
from openpyxl import load_workbook

# config
UPLOAD_FOLDER = 'uploads'
ALLOWED_IMG = {'png','jpg','jpeg','tif','tiff'}
ALLOWED_XLS = {'xlsx','xlsm','xls'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = os.environ.get('SECRET_KEY','devsecret')

def allowed_file(filename, allowed):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in allowed

def ocr_image(path):
    img = Image.open(path)
    gray = img.convert('L')
    gray = ImageOps.autocontrast(gray)
    gray = gray.filter(ImageFilter.MedianFilter(size=3))
    return pytesseract.image_to_string(gray, lang='eng')

def parse_ocr_text(text):
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    entries = []
    curr = {}
    for ln in lines:
        # league/title lines
        if re.search(r'League|Division|Veikkausliiga|V League|Pro League|Superliga|First Division', ln, re.I):
            if curr.get('home') and curr.get('away'):
                entries.append(curr)
                curr = {}
            curr['league'] = ln
            continue
        # time patterns
        if re.match(r'^\d{1,2}:\d{2}$', ln):
            curr['time'] = ln
            continue
        # Best Bet:
        mb = re.search(r'Best Bet[:\-]?\s*(.+)', ln, re.I)
        if mb:
            curr['best_bet'] = mb.group(1).strip()
            continue
        # vs line -> teams
        m = re.search(r'(.+?)\s+vs\.?\s+(.+)', ln, re.I)
        if m:
            curr['home'] = m.group(1).strip()
            curr['away'] = m.group(2).strip()
            continue
        # note (longer textual line)
        if len(ln.split()) > 2 and 'best' not in ln.lower():
            prev = curr.get('note','')
            curr['note'] = (prev + ' ' + ln).strip() if prev else ln
            continue
    if curr.get('home') and curr.get('away'):
        entries.append(curr)
    return entries

def excel_preview_html(excel_path, nrows=20):
    # Return HTML string to display first sheet as a table
    df = pd.read_excel(excel_path, sheet_name=0, nrows=nrows)
    return df.to_html(classes='table table-sm', index=False, na_rep='')

def find_header_row(ws):
    for r in range(1, ws.max_row+1):
        non_empty = sum(1 for c in ws[r] if c.value not in (None, ''))
        if non_empty >= 1:
            return r
    return 1

def map_headers(ws, header_row):
    headers = {}
    for col in range(1, ws.max_column+1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value:
            headers[str(cell.value).strip().lower()] = col
    return headers

def write_entries_to_template(template_path, entries, date_text, out_path):
    wb = load_workbook(template_path)
    ws = wb.active
    header_row = find_header_row(ws)
    header_map = map_headers(ws, header_row)

    # find first empty row after header_row
    write_row = header_row + 1
    # choose a column that is a known header: prefer 'home' or first header
    first_col = None
    for key in ['home','team','team name','match']:
        if key in header_map:
            first_col = header_map[key]
            break
    if first_col is None and header_map:
        first_col = next(iter(header_map.values()))
    # find first blank row in that column
    if first_col:
        r = header_row + 1
        while r <= ws.max_row and ws.cell(row=r, column=first_col).value not in (None, ''):
            r += 1
        write_row = r

    for e in entries:
        for h_lower, col in header_map.items():
            val = ''
            if 'home' in h_lower:
                val = e.get('home','')
            elif 'away' in h_lower or 'opponent' in h_lower:
                val = e.get('away','')
            elif 'league' in h_lower:
                val = e.get('league','')
            elif 'time' in h_lower:
                val = e.get('time','')
            elif 'date' in h_lower:
                try:
                    dt = datetime.fromisoformat(date_text)
                    val = dt.date()
                except Exception:
                    val = date_text
            elif 'bet' in h_lower or 'best' in h_lower:
                val = e.get('best_bet','')
            elif 'note' in h_lower or 'comment' in h_lower:
                val = e.get('note','')
            else:
                val = ''
            ws.cell(row=write_row, column=col).value = val
        write_row += 1

    wb.save(out_path)

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    template = request.files.get('template')
    image = request.files.get('image')
    if not template or not allowed_file(template.filename, ALLOWED_XLS):
        flash('Upload an Excel template (.xlsx).')
        return redirect(url_for('index'))
    if not image or not allowed_file(image.filename, ALLOWED_IMG):
        flash('Upload an image file (jpg/png).')
        return redirect(url_for('index'))

    tname = secure_filename(template.filename)
    iname = secure_filename(image.filename)
    template_path = os.path.join(app.config['UPLOAD_FOLDER'], 'template_' + tname)
    image_path = os.path.join(app.config['UPLOAD_FOLDER'], 'image_' + iname)
    template.save(template_path)
    image.save(image_path)

    # show preview (first rows)
    try:
        preview_html = excel_preview_html(template_path, nrows=20)
    except Exception as e:
        preview_html = f'<p>Error previewing Excel: {e}</p>'

    # OCR and parse immediately (show parsed preview)
    try:
        text = ocr_image(image_path)
        entries = parse_ocr_text(text)
    except Exception as e:
        flash('OCR failed: ' + str(e))
        return redirect(url_for('index'))

    request_id = f"{int(datetime.utcnow().timestamp())}"
    meta_path = os.path.join(app.config['UPLOAD_FOLDER'], f'meta_{request_id}.txt')
    with open(meta_path, 'w', encoding='utf8') as f:
        f.write(f"{template_path}\n{image_path}\n")

    df = pd.DataFrame(entries)
    parsed_html = df.to_html(classes='table table-sm', index=False, na_rep='')

    return render_template('preview.html',
                           preview_html=preview_html,
                           parsed_html=parsed_html,
                           request_id=request_id,
                           count=len(entries))

@app.route('/apply/<request_id>', methods=['POST'])
def apply(request_id):
    date_text = request.form.get('date', '')
    meta_path = os.path.join(app.config['UPLOAD_FOLDER'], f'meta_{request_id}.txt')
    if not os.path.exists(meta_path):
        flash('Session expired or files missing.')
        return redirect(url_for('index'))
    with open(meta_path, 'r', encoding='utf8') as f:
        template_path = f.readline().strip()
        image_path = f.readline().strip()

    text = ocr_image(image_path)
    entries = parse_ocr_text(text)
    out_path = os.path.join(app.config['UPLOAD_FOLDER'], f'filled_{request_id}.xlsx')
    write_entries_to_template(template_path, entries, date_text, out_path)
    return send_file(out_path, as_attachment=True, download_name='filled_predictions.xlsx')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
